import io
import time
from datetime import date

import folium
import openpyxl
import pandas as pd
import streamlit as st
from google.oauth2 import service_account
from google.oauth2.credentials import Credentials as UserCredentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from streamlit_folium import st_folium
from streamlit_js_eval import streamlit_js_eval

LOCATIONS_SHEET = "Locaties"
# Kolom-versies, nieuwste eerst. Oudere sheets (aangemaakt vóór een latere
# toevoeging zoals "Type" of de foto-kolommen) blijven zo leesbaar.
BASE_COLUMNS = ["Datum", "Naam", "Notities", "Latitude", "Longitude"]
COLUMNS_MET_TYPE = BASE_COLUMNS + ["Type"]
COLUMNS = COLUMNS_MET_TYPE + ["FotoId", "FotoLink"]
COLUMN_VERSIONS = [COLUMNS, COLUMNS_MET_TYPE, BASE_COLUMNS]
PHOTO_COLUMN = "G"  # enkel nog gebruikt voor foto's die al vóór de Drive-opslag waren ingebed
PHOTOS_FOLDER_NAME = "Leuke locaties - fotos"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
GOOGLE_SHEETS_MIME = "application/vnd.google-apps.spreadsheet"

TYPE_ICONS = {
    "Eten en drinken": "🍽️",
    "Bezichtigingen": "🏰",
    "Musea": "🖼️",
    "Slaaplocaties": "🛏️",
    "Wandelingen": "🥾",
    "Overig": "📍",
}
LOCATION_TYPES = list(TYPE_ICONS.keys())
DEFAULT_ICON = TYPE_ICONS["Overig"]


# Kleurenpalet van de banner, hergebruikt in alle andere styling zodat de rest van de
# app er als één geheel mee oogt i.p.v. een gekleurde banner boven een kale standaard-UI.
_DEEP_BLUE = "#0b3d63"
_TEAL = "#14919b"
_LIGHT_TEAL = "#45c2b3"


def inject_global_styles():
    st.markdown(
        f"""
        <style>
        .block-container {{
            padding-top: 2rem;
            max-width: 880px;
        }}
        [data-testid="stAppViewContainer"] {{
            background: linear-gradient(180deg, #f4fbfb 0%, #eef5f6 100%);
        }}
        hr {{
            height: 3px;
            border: none;
            border-radius: 999px;
            background: linear-gradient(90deg, {_DEEP_BLUE}, {_LIGHT_TEAL});
            opacity: 0.55;
        }}
        .leuke-section-header {{
            display: flex;
            align-items: center;
            gap: 0.6rem;
            margin: 1.8rem 0 1rem;
        }}
        .leuke-section-header .icon {{
            font-size: 1.5rem;
        }}
        .leuke-section-header h2 {{
            margin: 0;
            font-size: 1.4rem;
            font-weight: 700;
            color: {_DEEP_BLUE};
            white-space: nowrap;
        }}
        .leuke-section-header .bar {{
            flex: 1;
            height: 3px;
            border-radius: 999px;
            background: linear-gradient(90deg, {_TEAL}, transparent);
        }}
        .st-key-add_location_panel,
        .st-key-overview_panel,
        div[class*="st-key-card_"] {{
            border-radius: 16px !important;
            background: #ffffff;
            box-shadow: 0 4px 16px rgba(11, 61, 99, 0.08);
            transition: box-shadow 0.2s ease;
        }}
        div[class*="st-key-card_"]:hover {{
            box-shadow: 0 8px 22px rgba(11, 61, 99, 0.14);
        }}
        div[data-testid="stTextInput"] input,
        div[data-testid="stTextArea"] textarea,
        div[data-testid="stDateInput"] input,
        div[data-testid="stSelectbox"] div[data-baseweb="select"] {{
            border-radius: 10px !important;
        }}
        div[data-testid="stTextInput"] input:focus,
        div[data-testid="stTextArea"] textarea:focus {{
            border-color: {_TEAL} !important;
            box-shadow: 0 0 0 1px {_TEAL} !important;
        }}
        div[data-testid="stButton"] button,
        div[data-testid="stDownloadButton"] button {{
            border-radius: 999px;
            border: 1.5px solid {_TEAL};
            color: {_DEEP_BLUE};
            transition: all 0.15s ease;
        }}
        div[data-testid="stButton"] button:hover,
        div[data-testid="stDownloadButton"] button:hover {{
            border-color: {_DEEP_BLUE};
            color: #ffffff;
            background: linear-gradient(135deg, {_DEEP_BLUE}, {_TEAL});
        }}
        div[data-testid="stFormSubmitButton"] button {{
            border-radius: 999px;
            border: none;
            background: linear-gradient(135deg, {_DEEP_BLUE} 0%, {_TEAL} 100%);
            color: #ffffff;
            font-weight: 600;
            padding: 0.5rem 1.7rem;
            box-shadow: 0 4px 14px rgba(11, 61, 99, 0.25);
            transition: box-shadow 0.15s ease, transform 0.15s ease;
        }}
        div[data-testid="stFormSubmitButton"] button:hover {{
            box-shadow: 0 6px 18px rgba(11, 61, 99, 0.35);
            transform: translateY(-1px);
            color: #ffffff;
        }}
        div[data-testid="stTabs"] button[role="tab"] {{
            font-weight: 600;
            color: #4a6572;
        }}
        div[data-testid="stTabs"] button[aria-selected="true"] {{
            color: {_DEEP_BLUE};
        }}
        div[data-testid="stTabs"] [data-baseweb="tab-highlight"] {{
            background-color: {_TEAL} !important;
            height: 3px;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_banner():
    st.markdown(
        f"""
        <style>
        .leuke-banner {{
            background: linear-gradient(135deg, {_DEEP_BLUE} 0%, {_TEAL} 60%, {_LIGHT_TEAL} 100%);
            border-radius: 18px;
            padding: 2.2rem 2rem;
            margin-bottom: 1.5rem;
            box-shadow: 0 10px 30px rgba(11, 61, 99, 0.25);
            text-align: center;
        }}
        .leuke-banner h1 {{
            margin: 0;
            color: #ffffff;
            font-size: 2.6rem;
            font-weight: 800;
            letter-spacing: 0.02em;
            text-shadow: 0 2px 8px rgba(0, 0, 0, 0.15);
        }}
        .leuke-banner p {{
            margin: 0.4rem 0 0;
            color: rgba(255, 255, 255, 0.85);
            font-size: 1.05rem;
        }}
        </style>
        <div class="leuke-banner">
            <h1>\U0001F5FA️ Leuke locaties</h1>
            <p>Al jullie favoriete plekken, overzichtelijk verzameld.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_section_header(icon: str, text: str):
    st.markdown(
        f"""
        <div class="leuke-section-header">
            <span class="icon">{icon}</span>
            <h2>{text}</h2>
            <span class="bar"></span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def type_icon(type_locatie) -> str:
    return TYPE_ICONS.get(type_locatie, DEFAULT_ICON)


# Antwerpen, gebruikt als startpunt zolang er nog geen locatie gekozen is.
DEFAULT_LAT, DEFAULT_LON = 51.2194, 4.4025

# streamlit_js_eval's eigen get_geolocation() vraagt geen enableHighAccuracy aan,
# waardoor browsers vaak terugvallen op snelle maar grove wifi/cell-positionering
# (nauwkeurigheid van een hele buurt i.p.v. de exacte plek). Deze eigen variant
# vraagt expliciet GPS-nauwkeurigheid op, via dezelfde onderliggende JS-eval-component.
_GEOLOCATION_JS = """
new Promise((resolve) => {
    navigator.geolocation.getCurrentPosition(
        (position) => resolve({
            coords: {
                accuracy: position.coords.accuracy,
                altitude: position.coords.altitude,
                altitudeAccuracy: position.coords.altitudeAccuracy,
                heading: position.coords.heading,
                latitude: position.coords.latitude,
                longitude: position.coords.longitude,
                speed: position.coords.speed,
            },
            timestamp: position.timestamp,
        }),
        (error) => resolve({error: {code: error.code, message: error.message}}),
        {enableHighAccuracy: true, timeout: 15000, maximumAge: 0}
    );
})
"""


def get_precise_geolocation(component_key: str = "get_precise_location"):
    return streamlit_js_eval(js_expressions=_GEOLOCATION_JS, key=component_key)


def compress_photo(data: bytes, max_size: int = 1024, quality: int = 75) -> bytes:
    img = PILImage.open(io.BytesIO(data)).convert("RGB")
    img.thumbnail((max_size, max_size))
    out = io.BytesIO()
    img.save(out, format="JPEG", quality=quality)
    return out.getvalue()


def image_bytes(xl_image) -> bytes:
    try:
        return xl_image._data()
    except Exception:
        buf = io.BytesIO()
        xl_image.ref.save(buf, format="PNG")
        return buf.getvalue()


# --- Google Drive opslag ---
# Het bestand blijft altijd op Drive; de app haalt het op, wijzigt het in het
# geheugen en zet het meteen terug weg. Zo is er geen lokale schijf nodig en
# werkt dit ook vanaf de cloud.

@st.cache_resource
def get_drive_service():
    info = dict(st.secrets["gcp_service_account"])
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/drive"]
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _with_retries(func, attempts: int = 3, delay: float = 1.5):
    # Mobiele verbindingen vallen wel eens even weg (bv. SSL record layer failure).
    # Bij elke mislukte poging wordt ook de gecachte Drive-service weggegooid, voor
    # het geval een verouderde/kapotte verbinding in de cache het probleem was.
    last_exc = None
    for attempt in range(attempts):
        try:
            return func()
        except Exception as exc:
            last_exc = exc
            get_drive_service.clear()
            if attempt < attempts - 1:
                time.sleep(delay)
    raise last_exc


def download_workbook_bytes(file_id: str) -> bytes:
    def _do():
        service = get_drive_service()
        meta = service.files().get(fileId=file_id, fields="mimeType").execute()
        if meta["mimeType"] == GOOGLE_SHEETS_MIME:
            # Native Google Sheet (aangemaakt via "Nieuw -> Google Spreadsheets"):
            # moet geëxporteerd worden als xlsx, kan niet rechtstreeks gedownload worden.
            request = service.files().export_media(fileId=file_id, mimeType=XLSX_MIME)
        else:
            # Een echt geüpload .xlsx-bestand.
            request = service.files().get_media(fileId=file_id)
        buffer = io.BytesIO()
        downloader = MediaIoBaseDownload(buffer, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        return buffer.getvalue()

    return _with_retries(_do)


def upload_workbook_bytes(file_id: str, data: bytes) -> None:
    def _do():
        service = get_drive_service()
        media = MediaIoBaseUpload(io.BytesIO(data), mimetype=XLSX_MIME, resumable=False)
        service.files().update(fileId=file_id, media_body=media).execute()

    _with_retries(_do)


# --- Foto's op Drive, als jouzelf i.p.v. het service-account ---
# Service accounts hebben 0 bytes opslagquota en kunnen daarom geen nieuwe bestanden
# aanmaken (enkel bestaande, gedeelde bestanden lezen/bijwerken, zoals de sheet zelf).
# Voor foto's wordt daarom het toegangstoken van je eigen Google-login hergebruikt
# (st.user.tokens["access"], vereist expose_tokens + een Drive-scope in [auth]).

def get_user_drive_service():
    try:
        token = st.user.tokens["access"]
    except (AttributeError, KeyError) as exc:
        raise RuntimeError(
            "Geen Drive-toegangstoken beschikbaar. Controleer of 'expose_tokens' en de "
            "Drive-scope in de [auth]-secrets staan, en log opnieuw in."
        ) from exc
    creds = UserCredentials(token=token)
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _get_or_create_photos_folder(service) -> str:
    # Backslash en enkel aanhalingsteken moeten geëscaped worden in Drive's query-syntax,
    # anders breekt een naam die zo'n teken bevat de 'q'-parameter (zie ook: waarom de
    # foldernaam zelf geen apostrof meer heeft).
    veilige_naam = PHOTOS_FOLDER_NAME.replace("\\", "\\\\").replace("'", "\\'")
    query = (
        f"name = '{veilige_naam}' and "
        "mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    )
    result = service.files().list(q=query, fields="files(id)").execute()
    bestaande = result.get("files", [])
    if bestaande:
        return bestaande[0]["id"]
    folder = service.files().create(
        body={"name": PHOTOS_FOLDER_NAME, "mimeType": "application/vnd.google-apps.folder"},
        fields="id",
    ).execute()
    return folder["id"]


def upload_photo_to_drive(photo_bytes: bytes, filename: str) -> tuple[str, str]:
    def _do():
        service = get_user_drive_service()
        folder_id = _get_or_create_photos_folder(service)
        media = MediaIoBaseUpload(io.BytesIO(photo_bytes), mimetype="image/jpeg", resumable=False)
        file = service.files().create(
            body={"name": filename, "parents": [folder_id]},
            media_body=media,
            fields="id, webViewLink",
        ).execute()
        return file["id"], file["webViewLink"]

    return _with_retries(_do)


@st.cache_data(show_spinner=False)
def fetch_drive_photo_bytes(file_id: str) -> bytes | None:
    try:
        service = get_user_drive_service()
        request = service.files().get_media(fileId=file_id)
        buffer = io.BytesIO()
        downloader = MediaIoBaseDownload(buffer, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        return buffer.getvalue()
    except Exception:
        return None


def _embed_photo(ws, row_num: int, photo_bytes: bytes) -> None:
    pil_img = PILImage.open(io.BytesIO(photo_bytes))
    target_width = 160
    target_height = int(pil_img.height * (target_width / pil_img.width))
    xl_image = XLImage(io.BytesIO(photo_bytes))
    xl_image.width, xl_image.height = target_width, target_height
    ws.add_image(xl_image, f"{PHOTO_COLUMN}{row_num}")
    ws.row_dimensions[row_num].height = target_height * 0.75  # px -> punten


def _read_locations_sheet(ws):
    # Leest rechtstreeks via openpyxl i.p.v. pandas' header-gebaseerde inlezen: robuust
    # tegen een sheet waarvan de header-rij per ongeluk verdwenen/beschadigd is (bv. door
    # Google Sheets dat zijn eigen versie terugschrijft over een open bestand), en tegen
    # oudere sheets die nog niet alle kolommen (Type, FotoId, FotoLink) hebben.
    eerste_rij = [ws.cell(row=1, column=c + 1).value for c in range(len(COLUMNS))]
    kolommen = None
    for versie in COLUMN_VERSIONS:
        if eerste_rij[: len(versie)] == versie:
            kolommen = versie
            break
    eerste_datarij = 2 if kolommen is not None else 1
    if kolommen is None:
        kolommen = COLUMNS

    records = []
    sheet_rij_naar_index = {}
    for sheet_rij in range(eerste_datarij, ws.max_row + 1):
        values = [ws.cell(row=sheet_rij, column=c + 1).value for c in range(len(kolommen))]
        if all(v is None for v in values):
            continue
        sheet_rij_naar_index[sheet_rij] = len(records)
        record = dict(zip(kolommen, values))
        for col in COLUMNS:
            record.setdefault(col, "")
        records.append(record)

    photos_by_index = {}
    for xl_image in ws._images:
        sheet_rij = xl_image.anchor._from.row + 1  # 0-indexed -> 1-indexed
        if sheet_rij in sheet_rij_naar_index:
            photos_by_index[sheet_rij_naar_index[sheet_rij]] = image_bytes(xl_image)

    return records, photos_by_index


def save_location(file_id: str, row: dict, photo_bytes: bytes | None) -> None:
    # Foto eerst uploaden (als eigen gebruiker, die wél Drive-opslagquota heeft) zodat
    # we bij een uploadfout niets aan de sheet raken.
    foto_id = foto_link = ""
    if photo_bytes:
        bestandsnaam = f"{row['Naam']}_{row['Datum']}.jpg"
        foto_id, foto_link = upload_photo_to_drive(photo_bytes, bestandsnaam)

    data = download_workbook_bytes(file_id)
    wb = openpyxl.load_workbook(io.BytesIO(data))

    if LOCATIONS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(LOCATIONS_SHEET)
        ws.append(COLUMNS)
    else:
        ws = wb[LOCATIONS_SHEET]
        # Header aanvullen met eventueel ontbrekende (nieuwere) kolommen, zodat oudere
        # sheets meegroeien zonder bestaande koptekst te verstoren.
        for i, col in enumerate(COLUMNS):
            if ws.cell(row=1, column=i + 1).value in (None, ""):
                ws.cell(row=1, column=i + 1, value=col)

    volledige_row = {**row, "FotoId": foto_id, "FotoLink": foto_link}
    ws.append([volledige_row.get(col, "") for col in COLUMNS])

    out = io.BytesIO()
    wb.save(out)
    upload_workbook_bytes(file_id, out.getvalue())


def _rewrite_locations_sheet(file_id: str, mutate) -> None:
    # Blad volledig herbouwen i.p.v. rijen ter plekke te wijzigen/verwijderen: openpyxl
    # schuift afbeelding-ankers niet automatisch mee, dit voorkomt dat foto's bij de
    # verkeerde rij terechtkomen. `mutate(records, photos_by_index)` past beide lijsten
    # in place aan; een record op None zetten laat die rij gewoon weg bij het herbouwen.
    data = download_workbook_bytes(file_id)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    if LOCATIONS_SHEET not in wb.sheetnames:
        return

    records, photos_by_index = _read_locations_sheet(wb[LOCATIONS_SHEET])
    mutate(records, photos_by_index)

    del wb[LOCATIONS_SHEET]
    new_ws = wb.create_sheet(LOCATIONS_SHEET)
    new_ws.append(COLUMNS)
    for idx, record in enumerate(records):
        if record is None:
            continue
        new_ws.append([record.get(col, "") for col in COLUMNS])
        photo = photos_by_index.get(idx)
        if photo:
            _embed_photo(new_ws, new_ws.max_row, photo)

    out = io.BytesIO()
    wb.save(out)
    upload_workbook_bytes(file_id, out.getvalue())


def _delete_drive_photo(foto_id: str) -> None:
    if not foto_id:
        return
    try:
        get_user_drive_service().files().delete(fileId=foto_id).execute()
    except Exception:
        pass  # foto zelf verwijderen is best-effort, het record wordt sowieso aangepast


def delete_location(file_id: str, index_to_delete: int) -> None:
    def mutate(records, photos_by_index):
        if 0 <= index_to_delete < len(records):
            _delete_drive_photo(records[index_to_delete].get("FotoId"))
            records[index_to_delete] = None
            photos_by_index.pop(index_to_delete, None)

    _rewrite_locations_sheet(file_id, mutate)


def update_location(
    file_id: str, index_to_update: int, updates: dict,
    new_photo_bytes: bytes | None = None, remove_photo: bool = False,
) -> None:
    nieuwe_foto_id = nieuwe_foto_link = ""
    if new_photo_bytes:
        bestandsnaam = f"{updates.get('Naam', 'locatie')}_{updates.get('Datum', '')}.jpg"
        nieuwe_foto_id, nieuwe_foto_link = upload_photo_to_drive(new_photo_bytes, bestandsnaam)

    def mutate(records, photos_by_index):
        if not (0 <= index_to_update < len(records)):
            return
        oud_record = records[index_to_update]

        if new_photo_bytes or remove_photo:
            _delete_drive_photo(oud_record.get("FotoId"))
            photos_by_index.pop(index_to_update, None)  # oude ingebedde foto niet meenemen

        nieuw_record = {**oud_record, **updates}
        if new_photo_bytes:
            nieuw_record["FotoId"] = nieuwe_foto_id
            nieuw_record["FotoLink"] = nieuwe_foto_link
        elif remove_photo:
            nieuw_record["FotoId"] = ""
            nieuw_record["FotoLink"] = ""
        records[index_to_update] = nieuw_record

    _rewrite_locations_sheet(file_id, mutate)


def load_locations(file_id: str):
    try:
        data = download_workbook_bytes(file_id)
    except Exception as exc:
        st.error(f"Kon bestand niet ophalen van Google Drive: {exc}")
        return pd.DataFrame(columns=COLUMNS), {}

    wb = openpyxl.load_workbook(io.BytesIO(data))  # niet read_only: nodig om afbeeldingen te lezen
    if LOCATIONS_SHEET not in wb.sheetnames:
        return pd.DataFrame(columns=COLUMNS), {}

    records, row_photos = _read_locations_sheet(wb[LOCATIONS_SHEET])
    df = pd.DataFrame(records, columns=COLUMNS)
    return df, row_photos


st.set_page_config(page_title="Leuke locaties", page_icon="\U0001F4CD")
inject_global_styles()

# --- Configuratie-check: geeft een duidelijke melding i.p.v. een kale
# "Internal server error" zolang niet alle secrets zijn ingevuld. ---
REQUIRED_SECRETS = ["auth", "gcp_service_account", "drive_file_id"]
missing_secrets = [key for key in REQUIRED_SECRETS if key not in st.secrets]
if missing_secrets:
    render_banner()
    st.error(
        "De app is nog niet volledig geconfigureerd. Ontbrekende secrets: "
        + ", ".join(missing_secrets)
        + ". Voeg deze toe via Settings → Secrets op Streamlit Cloud."
    )
    st.stop()

# --- Login (enkel toegankelijk voor jezelf) ---
if not st.user.is_logged_in:
    render_banner()
    st.write("Log in met Google om je locaties te beheren.")
    st.button("Inloggen met Google", on_click=st.login)
    st.stop()

allowed_emails = st.secrets.get("allowed_emails", [])
if allowed_emails and st.user.email not in allowed_emails:
    st.error(f"Geen toegang voor {st.user.email}.")
    st.button("Uitloggen", on_click=st.logout)
    st.stop()

# --- Automatisch uitloggen na 30 minuten inactiviteit ---
# Streamlit voert het script enkel opnieuw uit bij interactie, dus dit controleert bij
# elke actie of het te lang geleden is sinds de vorige - niet een harde afsluiting op de
# seconde na precies 30 minuten als de app gewoon open blijft staan zonder iets te doen.
IDLE_TIMEOUT_SECONDS = 30 * 60
laatste_activiteit = st.session_state.get("laatste_activiteit")
if laatste_activiteit is not None and time.time() - laatste_activiteit > IDLE_TIMEOUT_SECONDS:
    st.session_state.pop("laatste_activiteit", None)
    st.info("Automatisch uitgelogd na 30 minuten inactiviteit.")
    st.logout()
    st.stop()
st.session_state.laatste_activiteit = time.time()

FILE_ID = st.secrets["drive_file_id"]

render_banner()
with st.sidebar:
    st.caption(f"Ingelogd als {st.user.email}")
    st.button("Uitloggen", on_click=st.logout)

render_section_header("➕", "Nieuwe locatie toevoegen")

if "lat" not in st.session_state:
    st.session_state.lat = DEFAULT_LAT
    st.session_state.lon = DEFAULT_LON

with st.container(border=True, key="add_location_panel"):
    geoloc = get_precise_geolocation()
    if st.button("📍 Gebruik mijn huidige locatie"):
        if geoloc and "coords" in geoloc:
            st.session_state.lat = geoloc["coords"]["latitude"]
            st.session_state.lon = geoloc["coords"]["longitude"]
            nauwkeurigheid = geoloc["coords"].get("accuracy")
            if nauwkeurigheid:
                st.caption(f"Nauwkeurigheid: ±{nauwkeurigheid:.0f} m")
        elif geoloc and "error" in geoloc:
            st.warning(
                f"Kon je locatie niet ophalen: {geoloc['error'].get('message', 'onbekende fout')}. "
                "Geef de browser toestemming voor locatietoegang en probeer opnieuw."
            )
        else:
            st.warning("Nog geen locatie beschikbaar, probeer het nog eens.")

    picker_map = folium.Map(location=[st.session_state.lat, st.session_state.lon], zoom_start=13)
    folium.Marker([st.session_state.lat, st.session_state.lon]).add_to(picker_map)
    map_click = st_folium(picker_map, height=350, use_container_width=True, key="location_picker")
    if map_click and map_click.get("last_clicked"):
        st.session_state.lat = map_click["last_clicked"]["lat"]
        st.session_state.lon = map_click["last_clicked"]["lng"]

    picker_maps_url = f"https://www.google.com/maps?q={st.session_state.lat},{st.session_state.lon}"
    st.caption(
        f"Geselecteerde locatie: {st.session_state.lat:.5f}, {st.session_state.lon:.5f} — "
        f"[📍 Bekijk in Google Maps]({picker_maps_url})"
    )

    # Buiten het formulier: widgets in een form reageren pas na "Toevoegen", waardoor
    # de camera/upload-widget hieronder anders niet meteen zou verschijnen/reageren.
    if "foto_reset" not in st.session_state:
        st.session_state.foto_reset = 0
    reset_suffix = st.session_state.foto_reset

    st.markdown("**Foto** (optioneel)")
    foto_modus = st.radio(
        "Foto", ["Geen", "Nemen", "Kiezen"], horizontal=True,
        key=f"foto_modus_{reset_suffix}", label_visibility="collapsed",
    )
    foto_ruw = None
    if foto_modus == "Nemen":
        camera_bestand = st.camera_input("Neem een foto", key=f"camera_foto_{reset_suffix}")
        if camera_bestand:
            foto_ruw = camera_bestand.getvalue()
    elif foto_modus == "Kiezen":
        upload_bestand = st.file_uploader(
            "Kies een foto", type=["jpg", "jpeg", "png"], key=f"upload_foto_{reset_suffix}"
        )
        if upload_bestand:
            foto_ruw = upload_bestand.getvalue()

    if foto_ruw:
        st.image(foto_ruw, caption="Voorbeeld", width=200)

    with st.form("nieuwe_locatie", clear_on_submit=True):
        naam = st.text_input("Naam")
        type_locatie = st.selectbox(
            "Type", LOCATION_TYPES, format_func=lambda t: f"{TYPE_ICONS[t]} {t}"
        )
        notities = st.text_area("Notities (optioneel)")
        datum = st.date_input("Datum", value=date.today())

        submitted = st.form_submit_button("Toevoegen")

        if submitted:
            if not naam:
                st.error("Vul een naam in.")
            else:
                row = {
                    "Datum": datum,
                    "Naam": naam,
                    "Type": type_locatie,
                    "Notities": notities,
                    "Latitude": st.session_state.lat,
                    "Longitude": st.session_state.lon,
                }
                foto_bytes = compress_photo(foto_ruw) if foto_ruw else None
                try:
                    save_location(FILE_ID, row, foto_bytes)
                except Exception as exc:
                    st.error(f"Kon niet opslaan naar Google Drive: {exc}")
                else:
                    st.success(f"'{naam}' toegevoegd!")
                    st.session_state.foto_reset += 1
                    st.rerun()

st.divider()
render_section_header("🌍", "Overzicht")

locaties, row_photos = load_locations(FILE_ID)

if locaties.empty:
    st.info("Nog geen locaties toegevoegd.")
else:
    with st.container(border=True, key="overview_panel"):
        zoek = st.text_input("Zoeken", placeholder="Naam of notities...")
        gefilterd = locaties
        if zoek:
            mask = (
                locaties["Naam"].str.contains(zoek, case=False, na=False)
                | locaties["Notities"].str.contains(zoek, case=False, na=False)
            )
            gefilterd = locaties[mask]

        gefilterd_geo = gefilterd.dropna(subset=["Latitude", "Longitude"])
        ontbrekend = len(gefilterd) - len(gefilterd_geo)
        if ontbrekend:
            st.warning(f"{ontbrekend} locatie(s) hebben geen geldige coördinaten en worden niet op de kaart getoond.")

        if not gefilterd_geo.empty:
            overview_map = folium.Map(
                location=[gefilterd_geo["Latitude"].mean(), gefilterd_geo["Longitude"].mean()],
                zoom_start=7,
            )
            for _, r in gefilterd_geo.iterrows():
                folium.Marker(
                    [r["Latitude"], r["Longitude"]],
                    popup=f"{type_icon(r['Type'])} {r['Naam']} ({r['Datum']})",
                    icon=folium.DivIcon(
                        html=f'<div style="font-size: 24px; transform: translate(-50%, -50%);">{type_icon(r["Type"])}</div>'
                    ),
                ).add_to(overview_map)
            st_folium(overview_map, height=350, use_container_width=True, key="overview_map")

    FOTO_BREEDTE = 120  # klein houden i.p.v. de volledige kolombreedte

    def render_location_card(idx, r, tab_key):
        card_key = f"card_{tab_key.replace(' ', '_')}_{idx}"
        with st.container(border=True, key=card_key):
            cols = st.columns([1, 2])
            with cols[0]:
                foto_id = r.get("FotoId")
                if foto_id:
                    foto_data = fetch_drive_photo_bytes(foto_id)
                    if foto_data:
                        st.image(foto_data, width=FOTO_BREEDTE)
                elif idx in row_photos:
                    st.image(row_photos[idx], width=FOTO_BREEDTE)
            with cols[1]:
                st.markdown(f"**{type_icon(r['Type'])} {r['Naam']}** — {r['Datum']}")
                if r.get("Notities"):
                    st.write(r["Notities"])
                if pd.notna(r["Latitude"]) and pd.notna(r["Longitude"]):
                    maps_url = f"https://www.google.com/maps?q={r['Latitude']},{r['Longitude']}"
                    st.markdown(f"[📍 Open in Google Maps]({maps_url})")
                if r.get("FotoLink"):
                    st.markdown(f"[🖼️ Open foto in Drive]({r['FotoLink']})")

                edit_key = f"editing_{tab_key}_{idx}"
                confirm_key = f"confirm_delete_{tab_key}_{idx}"

                if st.session_state.get(edit_key):
                    huidig_type = r["Type"] if r["Type"] in LOCATION_TYPES else LOCATION_TYPES[-1]
                    nieuwe_naam = st.text_input("Naam", value=r["Naam"], key=f"edit_naam_{tab_key}_{idx}")
                    nieuw_type = st.selectbox(
                        "Type", LOCATION_TYPES, index=LOCATION_TYPES.index(huidig_type),
                        format_func=lambda t: f"{TYPE_ICONS[t]} {t}", key=f"edit_type_{tab_key}_{idx}",
                    )
                    nieuwe_notities = st.text_area(
                        "Notities", value=r.get("Notities") or "", key=f"edit_notities_{tab_key}_{idx}"
                    )
                    nieuwe_datum = st.date_input("Datum", value=r["Datum"], key=f"edit_datum_{tab_key}_{idx}")
                    nieuwe_foto_bestand = st.file_uploader(
                        "Nieuwe foto (optioneel, vervangt de huidige)", type=["jpg", "jpeg", "png"],
                        key=f"edit_foto_{tab_key}_{idx}",
                    )
                    verwijder_foto = False
                    if r.get("FotoId") or idx in row_photos:
                        verwijder_foto = st.checkbox(
                            "Huidige foto verwijderen", key=f"edit_verwijder_foto_{tab_key}_{idx}"
                        )

                    ec1, ec2 = st.columns(2)
                    if ec1.button("Opslaan", key=f"edit_opslaan_{tab_key}_{idx}"):
                        updates = {
                            "Naam": nieuwe_naam,
                            "Type": nieuw_type,
                            "Notities": nieuwe_notities,
                            "Datum": nieuwe_datum,
                        }
                        nieuwe_foto_bytes = (
                            compress_photo(nieuwe_foto_bestand.getvalue()) if nieuwe_foto_bestand else None
                        )
                        try:
                            update_location(FILE_ID, idx, updates, nieuwe_foto_bytes, verwijder_foto)
                        except Exception as exc:
                            st.error(f"Kon niet bijwerken: {exc}")
                        else:
                            st.session_state.pop(edit_key, None)
                            st.success("Bijgewerkt.")
                            st.rerun()
                    if ec2.button("Annuleer", key=f"edit_annuleer_{tab_key}_{idx}"):
                        st.session_state.pop(edit_key, None)
                        st.rerun()

                elif st.session_state.get(confirm_key):
                    st.warning("Deze locatie definitief verwijderen?")
                    c1, c2 = st.columns(2)
                    if c1.button("Ja, verwijderen", key=f"yes_{tab_key}_{idx}"):
                        try:
                            delete_location(FILE_ID, idx)
                        except Exception as exc:
                            st.error(f"Kon niet verwijderen: {exc}")
                        else:
                            st.session_state.pop(confirm_key, None)
                            st.success("Verwijderd.")
                            st.rerun()
                    if c2.button("Annuleer", key=f"no_{tab_key}_{idx}"):
                        st.session_state.pop(confirm_key, None)
                        st.rerun()

                else:
                    bc1, bc2 = st.columns(2)
                    if bc1.button("✏️ Bewerken", key=f"bewerk_{tab_key}_{idx}"):
                        st.session_state[edit_key] = True
                        st.rerun()
                    if bc2.button("🗑️ Verwijderen", key=f"del_{tab_key}_{idx}"):
                        st.session_state[confirm_key] = True
                        st.rerun()

    tab_labels = ["Alle"] + [f"{TYPE_ICONS[t]} {t}" for t in LOCATION_TYPES]
    tabs = st.tabs(tab_labels)
    for tab, type_filter in zip(tabs, [None] + LOCATION_TYPES):
        with tab:
            subset = gefilterd if type_filter is None else gefilterd[gefilterd["Type"] == type_filter]
            if subset.empty:
                st.caption("Geen locaties in deze categorie.")
                continue
            tab_key = type_filter or "alle"
            for idx, r in subset.sort_values("Datum", ascending=False).iterrows():
                render_location_card(idx, r, tab_key)
